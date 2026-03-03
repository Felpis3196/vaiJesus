"""
Loader: carregamento de documentos (PDF, Excel, ODS, CSV), OCR, hyperlinks, load_document/load_data.
"""
import io
import os
import re
import pandas as pd
from datetime import datetime
from typing import Dict, List, Optional, Tuple, Union

from . import normalizer
from . import text_utils
from .normalizer import _try_normalize_pdf_mixed_text
from .text_utils import _parse_valor_monetario, _RE_VALOR_BR_COM_SINAL
from .ocr_preprocessor import extract_text_with_ocr, parse_ocr_text_to_dataframe, preprocess_scanned_pdf
from .normalizer import _infer_tipo_linha

# Limite de caracteres para extração de texto de PDF (balancete)
_LOAD_DOCUMENT_MAX_TEXT_LEN = 500_000
_BALANCETE_KEYWORDS = (
    "recebimentos", "despesas", "saldo inicial", "balancete", "ordinária", "ordinaria",
    "total dos recebimentos", "total das despesas", "recebimento no mês", "subtotal",
)


def _parse_ocr_text_to_transactions(text: str, ocr_used: bool = False) -> Optional[pd.DataFrame]:
    """
    Parse texto OCR de balancete e extrai transações estruturadas.
    Converte texto em DataFrame com colunas: data, descricao, valor, tipo.
    Usa a mesma lógica de _try_normalize_pdf_mixed_text mas adaptada para texto OCR.
    """
    if not text or len(text.strip()) < 100:
        return None
    
    import logging
    logger = logging.getLogger(__name__)
    
    # Usar função existente que já faz parsing de texto misto
    # Criar DataFrame temporário com texto na primeira coluna
    df_temp = pd.DataFrame({"texto": [text]})
    
    # Tentar usar função existente de normalização
    df_parsed = _try_normalize_pdf_mixed_text(df_temp)
    if df_parsed is not None and not df_parsed.empty:
        # Adicionar flag OCR
        df_parsed["_ocr_used"] = ocr_used
        logger.info(f"Extraídas {len(df_parsed)} transações do texto OCR usando parser existente")
        return df_parsed
    
    # Fallback: parsing manual mais simples
    transactions = []
    lines = text.split('\n')
    
    # Padrão para valores monetários (formato BR: 1.234,56 ou 1234,56)
    valor_pattern = re.compile(r'(-?\s*R\$\s*)?([\d]{1,3}(?:\.\d{3})*(?:,\d{2})?)(?=\s|$)', re.IGNORECASE)
    
    # Padrões para datas (DD/MM/YYYY, DD-MM-YYYY, DD.MM.YYYY)
    data_pattern = re.compile(r'(\d{1,2})[/\-\.](\d{1,2})[/\-\.](\d{2,4})')
    
    # Rastrear seção atual (receita vs despesa vs tabela explicativa de encargos)
    # Tabela explicativa NÃO entra na soma de débitos do resumo (apenas lançamentos reais)
    secao_atual = ""
    current_date = None
    # Apenas títulos de seção de tabela explicativa (não palavras em linhas de pagamento)
    explicativo_keywords = (
        "encargos trabalhistas", "base de cálculo", "base de calculo", "resumo encargos",
        "resumo encargo", "percentual", "encargos do período",
    )
    
    for line in lines:
        line = line.strip()
        if not line or len(line) < 5:
            continue
        
        line_lower = line.lower()
        
        # Seção de tabela explicativa (não usar como fonte de débitos do resumo)
        if any(k in line_lower for k in explicativo_keywords) and not data_pattern.search(line):
            secao_atual = "_explicativo"
            continue
        # Atualizar seção ao encontrar títulos de transações
        if "recebimentos" in line_lower or "ordinária" in line_lower or "ordinaria" in line_lower:
            secao_atual = "receita"
            continue
        if "despesas" in line_lower or "despesa " in line_lower:
            secao_atual = "despesa"
            continue
        
        # Tentar extrair data da linha
        data_match = data_pattern.search(line)
        if data_match:
            day, month, year = data_match.groups()
            year = int(year)
            if year < 100:
                year += 2000 if year < 50 else 1900
            try:
                current_date = f"{year:04d}-{month:02d}-{day:02d}"
            except:
                pass
        
        # Tentar extrair valor monetário usando padrão existente
        valor_matches = list(text_utils._RE_VALOR_BR_COM_SINAL.finditer(line))
        if not valor_matches:
            continue
        
        # Processar cada valor encontrado na linha
        for valor_match in valor_matches:
            valor_str = valor_match.group(0)
            is_neg = valor_str.startswith("-")
            valor = text_utils._parse_valor_monetario(valor_str.lstrip("-"))
            if is_neg:
                valor = -abs(valor)
            
            # Filtrar valores muito pequenos ou muito grandes (ruído do OCR)
            if abs(valor) < 0.01 or abs(valor) > 10_000_000:
                continue
            
            # Extrair descrição (texto antes do valor)
            descricao = line[:valor_match.start()].strip()
            # Remover data da descrição
            descricao = data_pattern.sub('', descricao).strip()
            # Limpar espaços múltiplos
            descricao = re.sub(r'\s+', ' ', descricao).strip()
            
            if len(descricao) < 2:
                descricao = "Item extraído do demonstrativo"
            
            # Classificar tipo usando função existente
            tipo = _infer_tipo_linha(line_lower, secao_atual)
            
            # Marcar linhas de tabela explicativa (encargos) para não somar como débito no resumo
            apenas_explicativo = secao_atual == "_explicativo"
            transactions.append({
                "data": current_date or datetime.now().strftime("%Y-%m-%d"),
                "descricao": descricao[:500],
                "valor": abs(valor) if valor < 0 else valor,  # Sempre positivo, tipo já indica receita/despesa
                "tipo": tipo,
                "_ocr_used": ocr_used,
                "_apenas_explicativo": apenas_explicativo,
            })
    
    if not transactions:
        logger.warning("Nenhuma transação extraída do texto OCR")
        return None
    
    df = pd.DataFrame(transactions)
    logger.info(f"Extraídas {len(df)} transações do texto OCR (parsing manual)")
    return df


def _try_ocr_pdf_text(pdf: "pdfplumber.PDF") -> tuple[str, bool]:
    """
    Tenta OCR nas páginas do PDF. Retorna (texto, ocr_used).
    Processa TODAS as páginas com resolução otimizada e configurações para português.
    """
    try:
        import pytesseract
        from PIL import Image
        t_cmd = os.getenv("TESSERACT_CMD")
        if t_cmd:
            pytesseract.pytesseract.tesseract_cmd = t_cmd
    except ImportError:
        return "", False
    except Exception:
        return "", False

    ocr_text_parts = []
    total_pages = len(pdf.pages)
    # Configurações otimizadas do Tesseract para português (com fallback)
    tesseract_configs = [
        '--oem 3 --psm 6 -l por',  # Tentar primeiro com português
        '--oem 3 --psm 6',  # Fallback sem idioma específico
    ]
    
    # Processar TODAS as páginas, não apenas 30
    config_used = None
    for page_num, page in enumerate(pdf.pages, 1):
        try:
            # Aumentar resolução para melhor qualidade (300 DPI)
            page_image = page.to_image(resolution=300).original
            if isinstance(page_image, Image.Image):
                # Pré-processar imagem para melhorar OCR
                if page_image.mode != 'RGB':
                    page_image = page_image.convert('RGB')
                
                # Tentar com diferentes configurações
                ocr_text = None
                for tesseract_config in tesseract_configs:
                    try:
                        ocr_text = pytesseract.image_to_string(page_image, config=tesseract_config)
                        if config_used is None:
                            config_used = tesseract_config
                        if ocr_text and len(ocr_text.strip()) > 10:
                            break  # Sucesso, usar esta configuração
                    except Exception:
                        continue  # Tentar próxima configuração
                
                if ocr_text and len(ocr_text.strip()) > 10:  # Ignorar textos muito pequenos
                    ocr_text_parts.append(ocr_text)
        except Exception as e:
            # Continuar processando outras páginas mesmo se uma falhar
            continue
    
    if ocr_text_parts:
        full_text = "\n".join(ocr_text_parts)
        return full_text, True
    return "", False


def _load_pdf_to_dataframe(source: Union[str, io.BytesIO]) -> pd.DataFrame:
    """
    Carrega um PDF em DataFrame. Para balancetes, prioriza TEXTO (até 500k caracteres)
    para não perder lançamentos; fallback em extract_tables().
    Detecta automaticamente PDFs escaneados e usa OCR quando necessário.
    source: caminho do arquivo (str) ou BytesIO com conteúdo do PDF.
    """
    import pdfplumber
    import logging
    logger = logging.getLogger(__name__)

    text = ""
    ocr_used = False
    with pdfplumber.open(source) as pdf:
        total_pages = len(pdf.pages)
        logger.info(f"[_load_pdf_to_dataframe] PDF tem {total_pages} páginas")
        
        # Tentar extrair texto de TODAS as páginas
        for page_num, page in enumerate(pdf.pages, 1):
            tx = page.extract_text()
            if tx and len(tx.strip()) > 10:  # Ignorar textos muito pequenos
                text += tx + "\n"
        
        # Detectar PDF escaneado: se extraiu muito pouco texto
        text_length = len(text.strip())
        chars_per_page = text_length / total_pages if total_pages > 0 else 0
        is_scanned = text_length < 100 or (total_pages > 0 and chars_per_page < 50)
        
        logger.info(f"[_load_pdf_to_dataframe] Texto extraído: {text_length} chars, média: {chars_per_page:.1f} chars/página, escaneado: {is_scanned}")
        
        # Se detectado como escaneado OU não extraiu texto, usar OCR automaticamente
        if is_scanned or not text:
            logger.info("[_load_pdf_to_dataframe] Tentando OCR...")
            ocr_text, ocr_used = _try_ocr_pdf_text(pdf)
            if ocr_text and len(ocr_text.strip()) > 100:
                text = ocr_text
                logger.info(f"[_load_pdf_to_dataframe] OCR extraiu {len(text)} caracteres")
            else:
                logger.warning(f"[_load_pdf_to_dataframe] OCR não extraiu texto suficiente ({len(ocr_text) if ocr_text else 0} chars)")
                # Tentar extrair tabelas mesmo sem texto
                logger.info("[_load_pdf_to_dataframe] Tentando extrair tabelas diretamente do PDF...")
                tables = []
                for page_num, page in enumerate(pdf.pages, 1):
                    try:
                        t = page.extract_tables()
                        if t:
                            tables.extend(t)
                            logger.debug(f"[_load_pdf_to_dataframe] Encontradas {len(t)} tabela(s) na página {page_num}")
                    except Exception as e:
                        logger.debug(f"[_load_pdf_to_dataframe] Erro ao extrair tabelas da página {page_num}: {e}")
                        continue
                
                if tables:
                    logger.info(f"[_load_pdf_to_dataframe] ✓ Extraídas {len(tables)} tabela(s) do PDF")
                    rows = []
                    for t in tables:
                        if not t:
                            continue
                        # Processar todas as linhas, incluindo cabeçalho se necessário
                        for row_idx, row in enumerate(t):
                            if row and any(cell is not None and str(cell).strip() for cell in row):
                                rows.append(row)
                    
                    if rows:
                        # Tentar identificar cabeçalho mais inteligentemente
                        # Se primeira linha parece cabeçalho (muitas strings, poucos números), usar como header
                        first_row = rows[0] if rows else []
                        is_header = False
                        if first_row:
                            str_count = sum(1 for c in first_row if c and isinstance(c, str) and not any(char.isdigit() for char in str(c)[:10]))
                            num_count = sum(1 for c in first_row if c and (isinstance(c, (int, float)) or (isinstance(c, str) and any(char.isdigit() for char in str(c)))))
                            is_header = str_count > num_count and str_count >= 2
                        
                        if is_header and len(rows) > 1:
                            h = rows[0]
                            data_rows = rows[1:]
                        else:
                            h = [f"col_{i}" for i in range(len(rows[0]))]
                            data_rows = rows
                        
                        ncols = len(h)
                        rows_ok = [list(r)[:ncols] + [None] * (ncols - len(r)) for r in data_rows[:2000]]  # Aumentado para 2000 linhas
                        df_tables = pd.DataFrame(rows_ok, columns=list(h))
                        logger.info(f"[_load_pdf_to_dataframe] ✓ DataFrame criado: {len(df_tables)} linhas, {len(df_tables.columns)} colunas")
                        return df_tables
    
    text = (text or "")[:_LOAD_DOCUMENT_MAX_TEXT_LEN].strip()

    is_balancete = len(text) > 400 and any(k in text.lower() for k in _BALANCETE_KEYWORDS)

    if is_balancete and text:
        # Tentar parsear texto OCR em transações estruturadas
        df_parsed = _parse_ocr_text_to_transactions(text, ocr_used)
        if df_parsed is not None and not df_parsed.empty:
            return df_parsed
        # Se parsing falhou mas temos texto, tentar extrair tabelas também
        logger.warning("Parsing de texto OCR falhou, tentando extrair tabelas do PDF...")
    
    # Se não tem texto suficiente OU parsing falhou, tentar extrair tabelas diretamente
    if not text or len(text.strip()) < 100:
        logger.info("[_load_pdf_to_dataframe] Texto insuficiente, tentando extração agressiva de tabelas...")
        tables = []
        if isinstance(source, io.BytesIO):
            source.seek(0)
        with pdfplumber.open(source) as pdf:
            # Processar TODAS as páginas procurando tabelas com múltiplas estratégias
            for page_num, page in enumerate(pdf.pages, 1):
                try:
                    # Estratégia 1: extract_tables() padrão
                    t = page.extract_tables()
                    if t:
                        tables.extend(t)
                        logger.debug(f"[_load_pdf_to_dataframe] Página {page_num}: Encontradas {len(t)} tabela(s) com extract_tables()")
                    
                    # Estratégia 2: Tentar detectar tabelas manualmente usando bounding boxes
                    try:
                        words = page.extract_words()
                        if words:
                            # Agrupar palavras por linhas e colunas (heurística simples)
                            # Isso pode ajudar quando extract_tables() falha
                            logger.debug(f"[_load_pdf_to_dataframe] Página {page_num}: {len(words)} palavras encontradas")
                    except Exception:
                        pass
                    
                except Exception as e:
                    logger.debug(f"[_load_pdf_to_dataframe] Erro ao extrair tabelas da página {page_num}: {e}")
                    continue
        
        if tables:
            logger.info(f"[_load_pdf_to_dataframe] ✓ Extraídas {len(tables)} tabela(s) do PDF (total)")
            rows = []
            for t_idx, t in enumerate(tables):
                if not t:
                    continue
                # Processar todas as linhas da tabela
                for row_idx, row in enumerate(t):
                    if row and any(cell is not None and str(cell).strip() for cell in row):
                        rows.append(row)
            
            if rows:
                # Tentar identificar cabeçalho mais inteligentemente
                first_row = rows[0] if rows else []
                is_header = False
                if first_row:
                    str_count = sum(1 for c in first_row if c and isinstance(c, str) and not any(char.isdigit() for char in str(c)[:10]))
                    num_count = sum(1 for c in first_row if c and (isinstance(c, (int, float)) or (isinstance(c, str) and any(char.isdigit() for char in str(c)))))
                    is_header = str_count > num_count and str_count >= 2
                
                if is_header and len(rows) > 1:
                    h = rows[0]
                    data_rows = rows[1:]
                    logger.debug(f"[_load_pdf_to_dataframe] Usando primeira linha como cabeçalho: {h[:5]}")
                else:
                    h = [f"col_{i}" for i in range(len(rows[0]))]
                    data_rows = rows
                
                ncols = len(h)
                rows_ok = [list(r)[:ncols] + [None] * (ncols - len(r)) for r in data_rows[:2000]]  # Aumentado para 2000 linhas
                df_tables = pd.DataFrame(rows_ok, columns=list(h))
                logger.info(f"[_load_pdf_to_dataframe] ✓ DataFrame criado a partir de tabelas: {len(df_tables)} linhas, {len(df_tables.columns)} colunas")
                return df_tables
            else:
                logger.warning("[_load_pdf_to_dataframe] Tabelas encontradas mas nenhuma linha válida extraída")
    
    # Se ainda não tem dados estruturados, retornar texto bruto
    if text and len(text.strip()) > 50:
        return pd.DataFrame({"texto_extraido": [text], "linha_numero": [1], "_ocr_used": [ocr_used], "_ocr_text_len": [len(text)]})
    
    # Último recurso: DataFrame vazio com mensagem
    return pd.DataFrame({
        "tipo_documento": ["PDF_ESCANEADO"],
        "mensagem": ["PDF sem texto extraível e sem tabelas detectadas. OCR pode ter falhado."],
        "_ocr_used": [ocr_used],
        "_ocr_text_len": [len(text) if text else 0],
    })

    # Se não tem texto, tentar extrair tabelas diretamente (mesmo sem OCR)
    if not text or len(text.strip()) < 50:
        import logging
        logger = logging.getLogger(__name__)
        logger.info("Texto insuficiente, tentando extrair tabelas diretamente do PDF (sem OCR)...")
        tables = []
        if isinstance(source, io.BytesIO):
            source.seek(0)
        with pdfplumber.open(source) as pdf:
            total_pages = len(pdf.pages)
            logger.info(f"Procurando tabelas em {total_pages} páginas...")
            # Processar TODAS as páginas procurando tabelas
            for page_num, page in enumerate(pdf.pages, 1):
                try:
                    t = page.extract_tables()
                    if t:
                        tables.extend(t)
                        logger.debug(f"Encontradas {len(t)} tabela(s) na página {page_num}")
                except Exception as e:
                    logger.debug(f"Erro ao extrair tabelas da página {page_num}: {e}")
                    continue
        
        if tables:
            logger.info(f"✓ Extraídas {len(tables)} tabela(s) do PDF")
            rows = []
            for t in tables:
                if not t:
                    continue
                for row in t[1:]:  # Pular cabeçalho
                    if row and any(cell is not None and str(cell).strip() for cell in row):
                        rows.append(row)
            
            if rows:
                h = tables[0][0] if tables[0] and tables[0][0] else [f"col_{i}" for i in range(len(rows[0]))]
                ncols = len(h)
                rows_ok = [list(r)[:ncols] + [None] * (ncols - len(r)) for r in rows[:1000]]  # Limitar a 1000 linhas
                df_tables = pd.DataFrame(rows_ok, columns=list(h))
                logger.info(f"✓ DataFrame criado a partir de tabelas: {len(df_tables)} linhas, {len(df_tables.columns)} colunas")
                return df_tables
            else:
                logger.warning("Tabelas encontradas mas sem dados válidos")
        else:
            logger.warning("Nenhuma tabela encontrada no PDF")
        
        return pd.DataFrame({
            "tipo_documento": ["PDF_ESCANEADO"],
            "mensagem": ["PDF sem texto extraível e sem tabelas detectadas. OCR pode ter falhado ou PDF está corrompido."],
            "_ocr_used": [ocr_used],
            "_ocr_text_len": [len(text) if text else 0],
        })

    # Texto existe mas não é balancete: usar tabelas se houver
    tables = []
    if isinstance(source, io.BytesIO):
        source.seek(0)
    with pdfplumber.open(source) as pdf:
        # Processar TODAS as páginas procurando tabelas
        for page in pdf.pages:
            t = page.extract_tables()
            if t:
                tables.extend(t)
    if tables:
        rows = []
        for t in tables:
            if not t:
                continue
            for row in t[1:]:
                if row and any(cell is not None for cell in row):
                    rows.append(row)
        if rows:
            h = tables[0][0] if tables[0] else [f"col_{i}" for i in range(len(rows[0]))]
            ncols = len(h)
            rows_ok = [list(r)[:ncols] + [None] * (ncols - len(r)) for r in rows[:500]]
            return pd.DataFrame(rows_ok, columns=list(h))
    
    # Se tem texto OCR mas não encontrou tabelas, tentar parsear texto diretamente
    if ocr_used and text and len(text.strip()) > 100:
        df_parsed = _parse_ocr_text_to_transactions(text, ocr_used)
        if df_parsed is not None and not df_parsed.empty:
            return df_parsed
    
    return pd.DataFrame({"texto_extraido": [text], "linha_numero": [1], "_ocr_used": [ocr_used], "_ocr_text_len": [len(text)]})


def _get_pdf_text_and_page_info(file_path: str) -> Tuple[str, Optional[Dict[int, int]]]:
    """
    Extrai texto do PDF uma única vez: pdfplumber (digital) ou OCR (escaneado).
    Retorna (texto, page_info) onde page_info é None para PDF digital.
    """
    import logging
    logger = logging.getLogger(__name__)
    text = ""
    try:
        import pdfplumber
        with pdfplumber.open(file_path) as pdf:
            total_pages = len(pdf.pages)
            for page_num, page in enumerate(pdf.pages, 1):
                tx = page.extract_text()
                if tx and len(tx.strip()) > 10:
                    text += tx + "\n"
            text_length = len(text.strip())
            chars_per_page = text_length / total_pages if total_pages > 0 else 0
            is_scanned = text_length < 100 or (total_pages > 0 and chars_per_page < 50)
            if not is_scanned and text_length >= 100:
                logger.info(f"[_get_pdf_text] PDF digital: {text_length} caracteres")
                return (text.strip(), None)
        # PDF escaneado ou pouco texto: usar OCR (todas as páginas)
        # from .ocr_preprocessor (top)
        ocr_text, ocr_used, page_info = extract_text_with_ocr(
            file_path, max_pages=None, return_page_info=True
        )
        if ocr_text and len(ocr_text.strip()) > 0:
            logger.info(f"[_get_pdf_text] OCR: {len(ocr_text)} caracteres")
            return (ocr_text.strip(), page_info if ocr_used else None)
        if text and len(text.strip()) > 0:
            return (text.strip(), None)
    except Exception as e:
        logger.warning(f"[_get_pdf_text] Erro: {e}")
    return ("", None)


# Palavras-chave que classificam uma URL como link de holerite/FGTS (tipo "fgts_holerite").
# Qualquer URL cujo texto (lower-case) contiver ao menos uma dessas palavras recebe prioridade
# no pipeline de extração de holerites.
_HOLERITE_URL_KEYWORDS = (
    "fgts", "holerite", "caixa.gov", "meu.inss",
    "folha", "contracheque", "contra-cheque", "payroll",
    "salario", "sal%C3%A1rio",  # URL-encoded "salário"
    "recibo", "pagamento", "demonstrativo",
)


def extract_hyperlinks_from_excel(file_path: str) -> List[Dict[str, str]]:
    """
    Extrai hyperlinks de arquivos Excel (.xls, .xlsx).
    Retorna lista de {"tipo": "fgts_holerite"|"outro", "url": str, "celula": "A1"}.
    """
    if not file_path or not os.path.exists(file_path):
        return []
    ext = os.path.splitext(file_path)[1].lower()
    links: List[Dict[str, str]] = []
    try:
        if ext in [".xls", ".xlt"]:
            try:
                import xlrd
                book = xlrd.open_workbook(file_path, formatting_info=True)
                for sheet_idx in range(book.nsheets):
                    sheet = book.sheet_by_index(sheet_idx)
                    if not hasattr(sheet, "hyperlink_map"):
                        continue
                    for (row_idx, col_idx), hl in sheet.hyperlink_map.items():
                        url = getattr(hl, "url_or_path", None) or getattr(hl, "url", None) or ""
                        if url and isinstance(url, str) and (url.startswith("http") or url.startswith("mailto:")):
                            def _col_letter(c: int) -> str:
                                s = ""
                                while c >= 0:
                                    s = chr(c % 26 + 65) + s
                                    c = c // 26 - 1
                                return s or "A"
                            col_letter = _col_letter(col_idx) if col_idx is not None else ""
                            ref = f"{col_letter}{row_idx + 1}"
                            tipo = "fgts_holerite" if any(x in url.lower() for x in _HOLERITE_URL_KEYWORDS) else "outro"
                            links.append({"tipo": tipo, "url": url.strip(), "celula": ref})
            except Exception as e:
                import logging
                logging.getLogger(__name__).debug(f"xlrd hyperlinks: {e}")
        elif ext == ".xlsx":
            try:
                import openpyxl
                wb = openpyxl.load_workbook(file_path, read_only=False, data_only=True)
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    for row in ws.iter_rows():
                        for cell in row:
                            if cell.hyperlink:
                                target = getattr(cell.hyperlink, "target", None) or str(cell.hyperlink)
                                if target and isinstance(target, str) and (target.startswith("http") or target.startswith("mailto:")):
                                    ref = cell.coordinate
                                    tipo = "fgts_holerite" if any(x in target.lower() for x in _HOLERITE_URL_KEYWORDS) else "outro"
                                    links.append({"tipo": tipo, "url": target.strip(), "celula": ref})
                wb.close()
            except Exception as e:
                import logging
                logging.getLogger(__name__).debug(f"openpyxl hyperlinks: {e}")
    except Exception as e:
        import logging
        logging.getLogger(__name__).warning(f"extract_hyperlinks_from_excel: {e}")
    return links


def extract_hyperlinks_from_ods(file_path: str) -> List[Dict[str, str]]:
    """
    Extrai hyperlinks de arquivos ODS usando odfpy.
    Retorna lista de {"tipo": "fgts_holerite"|"outro", "url": str, "celula": ref}.
    """
    if not file_path or not os.path.exists(file_path):
        return []
    if os.path.splitext(file_path)[1].lower() != ".ods":
        return []
    links: List[Dict[str, str]] = []
    try:
        from odf.opendocument import load
        from odf import text as odf_text
        doc = load(file_path)
        if not hasattr(doc, "spreadsheet") or doc.spreadsheet is None:
            return []
        count = 0
        for elem in doc.spreadsheet.getElementsByType(odf_text.A):
            href = elem.getAttribute("href") if hasattr(elem, "getAttribute") else None
            if not href or not isinstance(href, str) or not (href.startswith("http") or href.startswith("mailto:")):
                continue
            count += 1
            ref = f"L{count}"
            tipo = "fgts_holerite" if any(x in href.lower() for x in _HOLERITE_URL_KEYWORDS) else "outro"
            links.append({"tipo": tipo, "url": href.strip(), "celula": ref})
    except Exception as e:
        import logging
        logging.getLogger(__name__).debug(f"extract_hyperlinks_from_ods: {e}")
    return links


def _detect_balancete_header_from_df(df: pd.DataFrame, max_rows: int = 25) -> Optional[int]:
    """
    Detecta a linha do DataFrame que contém o cabeçalho de balancete (Conta, Débitos, Créditos).
    Retorna o índice da linha ou None se não encontrar.
    """
    if df.empty or len(df) < 2:
        return None
    for idx in range(min(len(df), max_rows)):
        row = df.iloc[idx]
        row_str = " ".join(str(x).lower().replace("é", "e").replace("ê", "e") for x in row if pd.notna(x) and str(x).strip())
        if "conta" in row_str and (
            "debitos" in row_str or "debito" in row_str or "creditos" in row_str or "credito" in row_str
        ):
            return idx
    return None


def _detect_excel_balancete_header(file_path: str, engine: str = "xlrd", max_rows: int = 25) -> Optional[int]:
    """
    Detecta a linha que contém o cabeçalho de balancete (Conta, Débitos, Créditos).
    Retorna o índice da linha ou None se não encontrar.
    """
    try:
        df = pd.read_excel(file_path, engine=engine, header=None, nrows=max_rows)
        return _detect_balancete_header_from_df(df, max_rows)
    except Exception:
        pass
    return None


def load_document(file_path: str) -> Union[pd.DataFrame, Tuple[pd.DataFrame, str]]:
    """
    Carregador unificado: CSV, Excel, ODS e PDF.
    Para PDF tipo balancete, prioriza texto (até 500k caracteres); mesmo comportamento
    em todos os fluxos (script, API, DataInputManager, main).
    Para PDFs escaneados, usa pré-processamento OCR dedicado.
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Arquivo não encontrado: {file_path}")

    ext = os.path.splitext(file_path)[1].lower()
    if ext in [".ods", ".sxc"]:
        try:
            return pd.read_excel(file_path, engine="odf")
        except Exception as e:
            if ext == ".sxc":
                # Alguns .sxc podem ser zip/XML; tentar ler como Excel antigo (xlrd) se odf falhar
                try:
                    return pd.read_excel(file_path, engine="xlrd")
                except Exception:
                    raise e
            raise
    if ext in [".xlsx", ".xls", ".xlt"]:
        # .xlt e .xls são formato Excel 97-2003; precisam do engine xlrd
        if ext in [".xls", ".xlt"]:
            try:
                # Detectar linha de cabeçalho com Conta + Débitos/Créditos (prestação de contas)
                header_row = _detect_excel_balancete_header(file_path, engine="xlrd")
                if header_row is not None:
                    return pd.read_excel(file_path, engine="xlrd", header=header_row)
                return pd.read_excel(file_path, engine="xlrd")
            except Exception as e:
                if ext == ".xlt":
                    try:
                        return pd.read_excel(file_path, engine="odf")
                    except Exception:
                        raise e
                raise
        return pd.read_excel(file_path)
    if ext == ".csv":
        return pd.read_csv(file_path, encoding="utf-8")
    if ext == ".pdf":
        import logging
        logger = logging.getLogger(__name__)
        logger.info(f"[LOAD_DOCUMENT] ========== PROCESSANDO PDF ==========")
        logger.info(f"[LOAD_DOCUMENT] Arquivo: {file_path}")
        try:
            file_size = os.path.getsize(file_path)
            logger.info(f"[LOAD_DOCUMENT] Tamanho: {file_size / 1024 / 1024:.2f} MB")
        except Exception:
            pass

        # Uma única extração de texto (pdfplumber ou OCR, todas as páginas)
        text, page_info = _get_pdf_text_and_page_info(file_path)

        if text and len(text.strip()) > 0:
            # from .ocr_preprocessor (top)
            df = parse_ocr_text_to_dataframe(text, page_info=page_info)
            if df is not None and not df.empty:
                df["_ocr_used"] = bool(page_info)
                df["_source_file"] = os.path.basename(file_path)
                logger.info(f"[LOAD_DOCUMENT] ✓ {len(df)} transações extraídas do texto único")
                if "valor" in df.columns:
                    total_valor = df["valor"].sum()
                    logger.info(f"[LOAD_DOCUMENT] Total valores: R$ {total_valor:,.2f}")
                return (df, text)

            # Parsing falhou: fallback para método padrão
            logger.warning("[LOAD_DOCUMENT] Parsing retornou vazio, usando método padrão...")
            df_std = _load_pdf_to_dataframe(file_path)
            logger.info(f"[LOAD_DOCUMENT] Método padrão: {len(df_std)} linhas")
            return (df_std, text)

        # Sem texto: tentar preprocess_scanned_pdf e depois _load_pdf_to_dataframe (retorno só df)
        try:
            # from .ocr_preprocessor (top)
            df_ocr = preprocess_scanned_pdf(file_path)
            if df_ocr is not None and not df_ocr.empty:
                logger.info(f"[LOAD_DOCUMENT] ✓ Método completo: {len(df_ocr)} transações")
                return df_ocr
        except Exception as e:
            logger.debug(f"[LOAD_DOCUMENT] preprocess_scanned_pdf: {e}")

        logger.info("[LOAD_DOCUMENT] Usando método padrão de carregamento de PDF...")
        df_std = _load_pdf_to_dataframe(file_path)
        logger.info(f"[LOAD_DOCUMENT] Método padrão retornou {len(df_std)} linhas")
        return df_std
    raise ValueError(f"Formato não suportado: {ext}. Use .csv, .xls, .xlt, .xlsx, .ods, .sxc ou .pdf.")


def load_document_from_bytes(content: bytes, filename: str) -> pd.DataFrame:
    """
    Carrega documento a partir de bytes (ex.: upload na API).
    filename é usado para detectar extensão (.csv, .xlsx, .xls, .xlt, .ods, .sxc, .pdf).
    Para PDFs escaneados, salva temporariamente e usa pré-processamento OCR.
    """
    ext = os.path.splitext(filename)[1].lower() if filename else ""
    if ext in [".ods", ".sxc"]:
        try:
            return pd.read_excel(io.BytesIO(content), engine="odf")
        except Exception as e:
            if ext == ".sxc":
                try:
                    return pd.read_excel(io.BytesIO(content), engine="xlrd")
                except Exception:
                    raise e
            raise
    if ext in [".xlsx", ".xls", ".xlt"]:
        if ext in [".xls", ".xlt"]:
            try:
                # Detectar header de balancete (Conta + Débitos/Créditos) também ao carregar por bytes
                head_df = pd.read_excel(io.BytesIO(content), engine="xlrd", header=None, nrows=25)
                header_row = _detect_balancete_header_from_df(head_df, max_rows=25)
                if header_row is not None:
                    return pd.read_excel(io.BytesIO(content), engine="xlrd", header=header_row)
                return pd.read_excel(io.BytesIO(content), engine="xlrd")
            except Exception as e:
                if ext == ".xlt":
                    try:
                        return pd.read_excel(io.BytesIO(content), engine="odf")
                    except Exception:
                        raise e
                raise
        return pd.read_excel(io.BytesIO(content))
    if ext == ".csv":
        return pd.read_csv(io.BytesIO(content), encoding="utf-8")
    if ext == ".pdf":
        # Salvar temporariamente e usar load_document (retorna (df, text) ou df)
        try:
            import tempfile
            import logging
            logger = logging.getLogger(__name__)
            with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp_file:
                tmp_path = tmp_file.name
                tmp_file.write(content)
            try:
                result = load_document(tmp_path)
                df = result[0] if isinstance(result, tuple) else result
                return df
            finally:
                try:
                    os.unlink(tmp_path)
                except Exception:
                    pass
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.warning(f"Carregamento PDF de bytes falhou: {e}")
        return _load_pdf_to_dataframe(io.BytesIO(content))
    raise ValueError(f"Formato não suportado: {ext}. Use .csv, .xls, .xlt, .xlsx, .ods, .sxc ou .pdf.")


def load_data(file_path: str) -> pd.DataFrame:
    """Carrega dados financeiros de um arquivo (CSV, Excel, ODS ou PDF)."""
    result = load_document(file_path)
    return result[0] if isinstance(result, tuple) else result
